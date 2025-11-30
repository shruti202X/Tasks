import argparse
import csv
from datetime import datetime
import os
import pandas as pd
from utility import *
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import PatternFill, Alignment, Font # type: ignore
from openpyxl.styles.borders import Border, Side # type: ignore

to_choices = ["start", "rant", "pause", "complete", "list", "checkpoint", "merge", "view", "archive", "de-archive"]

parser = argparse.ArgumentParser(description="Better Manage Tasks")
parser.add_argument("--task", type=str, help="Name of the task or subtask", default=None)
parser.add_argument("--to", type=str, choices=to_choices, help="Specify action you want to perform.", default=None)
parser.add_argument("--extension", type=str, help="Enter extension for storing task files. (Default value is '.xlsx' for Excel files.)", default=".xlsx")
args = parser.parse_args()

task = args.task
if task is None:
    task = input("Enter Task: ")
extension = args.extension

task_file_name = task + extension

to = args.to
if to not in to_choices:
    to = input("Enter message tag: ")

while to not in to_choices:
    print("------------------------------------------------------------------")
    print("Message Tag: What does your message want to express?")
    print("- 'list': To describe or list down the task.")
    print("- 'start': Starting of the task, after a break.")
    print("- 'rant': Rant about the task.")
    print("- 'checkpoint': Completion of a sub task.")
    print("- 'pause': A short or long pause in the task.")
    print("- 'complete': Completion of the task.")
    print("- 'merge': To merge it with other tasks.")
    print("- 'archive': To archive the task.  (No message is added here!)")
    print("- 'de-archive': To move back the archived task to active task. (No message is added here!)")
    print("- 'view': To view/display the task. (No message is added here!)")
    print("------------------------------------------------------------------")
    to = input("Enter message tag: ")

if to == 'archive':
    os.makedirs("archive", exist_ok=True)
    if(not os.path.exists(task_file_name)):
        print("No such task found.")
    archive_task_file_name = os.path.join("archive", task_file_name)
    if (os.path.exists(archive_task_file_name)):
        merge_req = input("This task is already there in archive!\nReply with 'y' to merge and 'o' to over-write: ")
        if merge_req.startswith("y") or merge_req.startswith("Y"):
            merge_tasks(archive_task_file_name, set([task_file_name]))
            exit()
        elif merge_req.startswith("o") or merge_req.startswith("O"):
            os.remove(archive_task_file_name)
        else:
            print("You neither entered 'y', nor 'o'. Hence, stopping this run.")
            exit()
    os.rename(task_file_name, archive_task_file_name)
    exit()

if to == 'de-archive':
    archive_task_file_name = os.path.join("archive", task_file_name)
    if(not os.path.exists(archive_task_file_name)):
        print("No such archived task found.")
    if (os.path.exists(task_file_name)):
        merge_req = input("This task is already in active!\nReply with 'y' to merge and 'o' to over-write: ")
        if merge_req.startswith("y") or merge_req.startswith("Y"):
            merge_tasks(task_file_name, set([archive_task_file_name]))
            exit()
        elif merge_req.startswith("o") or merge_req.startswith("O"):
            os.remove(task_file_name)
        else:
            print("You neither entered 'y', nor 'o'. Hence, stopping this run.")
            exit()
    os.rename(archive_task_file_name, task_file_name)
    exit()

def format_excel(task_file_name):
    wb = load_workbook(task_file_name)
    df = pd.read_excel(task_file_name)
    align = Alignment(wrap_text=True,vertical='top')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for ws in wb.worksheets:
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 90
    for i in range(2, df.shape[0]+2):
        cell = ws.cell(row=i, column=4)
        cell.alignment = align
    for i in range(2, df.shape[0]+2):
        cell_1 = ws.cell(row=i, column=1)
        tag = cell_1.value
        if tag == "list":
            color = PatternFill(patternType='solid', fgColor='FFFFE0')
        elif tag == "start":
            color = PatternFill(patternType='solid', fgColor="A4FCED")
        elif tag == "rant":
            color = PatternFill(patternType='solid', fgColor='FFB6C1')
        elif tag == "checkpoint":
            color = PatternFill(patternType='solid', fgColor="B2F996")
        elif tag == "pause":
            color = PatternFill(patternType='solid', fgColor="FF9F90")
        elif tag == "complete":
            color = PatternFill(patternType='solid', fgColor="EC96D3")
        elif tag == "merge":
            color = PatternFill(patternType='solid', fgColor="BEC2F5")
        else:
            color = PatternFill(patternType='solid', fgColor='FFFFFF')
        cell_1.fill = color
        cell_1.border = thin_border
        cell_2 = ws.cell(row=i, column=2)
        cell_2.fill = color
        cell_2.border = thin_border
        cell_3 = ws.cell(row=i, column=3)
        cell_3.fill = color
        cell_3.border = thin_border
        cell_4 = ws.cell(row=i, column=4)
        cell_4.fill = color
        cell_4.border = thin_border
    wb.save(task_file_name)

if to == 'view':
    try:
        format_excel(task_file_name)
    except Exception as e:
        print(f"WARNING: Encountered Exception while formatting: {e}")
    try:
        os.popen(f'start excel.exe {task_file_name}')
    except Exception as e:
        print(f"ERROR: Unable to open task file: {task_file_name}.")
    exit()

if to == 'merge':
    other_tasks = set()
    other_task = input("Enter other task: ")
    while other_task!="":
        if os.path.exists(other_task+extension) == False:
           print("No such active task exists.")
        else:
            other_tasks.add(other_task+extension)
        other_task = input("Enter another task to add, else simply press enter: ")
    merge_tasks(task_file_name, other_tasks)

print("Enter message (use Ctrl+D to end):")
message = []
while True:
    message_line = input()
    if '\x04' in message_line: # If Ctrl D in line
        message.append(message_line.split('\x04')[0])
        break
    else:
        message.append(message_line)

message_string = '\n'.join(message)

heading = ["Tag", "Task/Subtask", "Date Time", "Message"]

if extension == ".xlsx":
    df = pd.DataFrame([[to, task, datetime.now().strftime(f"%d %B, %Y (%I:%M %p)"), message_string]], columns=heading)
    if os.path.exists(task_file_name):
        df2 = pd.read_excel(task_file_name)
        df = pd.concat([df, pd.read_excel(task_file_name)])
    with pd.ExcelWriter(task_file_name) as writer:
        df.to_excel(writer, sheet_name="Log", index=False)

else:
    with open(task_file_name, mode="a") as f:
        writer = csv.writer(f)
        if os.path.exists(task_file_name) == False:
            writer.writerow(heading)
        writer.writerow([to, task, datetime.now().strftime(f"%d %B, %Y (%I:%M %p)"), message_string])
